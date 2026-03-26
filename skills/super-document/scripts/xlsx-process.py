#!/usr/bin/env python3
"""
Excel 表格处理工具
使用 openpyxl 和 pandas 进行 Excel 读写、格式化、数据分析
"""

import argparse
import sys
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    print("请安装 pandas: pip install pandas")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
except ImportError:
    print("请安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# 样式预设
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_workbook() -> Workbook:
    """创建新工作簿"""
    wb = Workbook()
    return wb


def read_excel(file_path: str, sheet: str = None, header: int = 0) -> pd.DataFrame:
    """读取 Excel 文件"""
    df = pd.read_excel(file_path, sheet_name=sheet, header=header)
    return df


def write_excel(df: pd.DataFrame, output: str, sheet: str = "Sheet1",
                header_style: bool = True, index: bool = False):
    """写入 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            
            # 表头样式
            if r_idx == 1 and header_style:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
            else:
                cell.alignment = DATA_ALIGNMENT
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    print(f"✓ Excel 已保存到: {output}")


def apply_formatting(wb: Workbook, sheet_name: str, format_rules: dict):
    """应用格式化"""
    ws = wb[sheet_name]
    
    for rule in format_rules.get("rules", []):
        rule_type = rule.get("type")
        cells = rule.get("cells", "A1:Z100")
        
        if rule_type == "font":
            font = Font(**rule.get("options", {}))
            for row in ws[cells]:
                for cell in row:
                    cell.font = font
        
        elif rule_type == "fill":
            fill = PatternFill(**rule.get("options", {}))
            for row in ws[cells]:
                for cell in row:
                    cell.fill = fill
        
        elif rule_type == "alignment":
            alignment = Alignment(**rule.get("options", {}))
            for row in ws[cells]:
                for cell in row:
                    cell.alignment = alignment
        
        elif rule_type == "border":
            border = Border(**rule.get("options", {}))
            for row in ws[cells]:
                for cell in row:
                    cell.border = border


def add_formula(ws, cell: str, formula: str):
    """添加公式"""
    ws[cell] = formula


def add_chart(wb: Workbook, sheet_name: str, chart_type: str, 
              data_range: str, title: str = None, position: str = "E2"):
    """添加图表"""
    ws = wb[sheet_name]
    
    # 解析数据范围
    parts = data_range.split(":")
    start = parts[0]
    end = parts[1] if len(parts) > 1 else parts[0]
    
    # 创建图表
    if chart_type == "bar":
        chart = BarChart()
    elif chart_type == "pie":
        chart = PieChart()
    elif chart_type == "line":
        chart = LineChart()
    else:
        chart = BarChart()
    
    # 设置数据
    data = Reference(ws, range_string=f"{sheet_name}!{data_range}")
    chart.add_data(data)
    
    if title:
        chart.title = title
    
    ws.add_chart(chart, position)


def analyze_data(file_path: str, sheet: str = None):
    """数据分析"""
    df = read_excel(file_path, sheet)
    
    print("\n=== 数据概览 ===")
    print(f"行数: {len(df)}")
    print(f"列数: {len(df.columns)}")
    print(f"列名: {', '.join(df.columns.tolist())}")
    
    print("\n=== 数据类型 ===")
    print(df.dtypes)
    
    print("\n=== 数值列统计 ===")
    print(df.describe())
    
    print("\n=== 缺失值 ===")
    print(df.isnull().sum())
    
    return df


def filter_data(file_path: str, column: str, value, output: str):
    """筛选数据"""
    df = read_excel(file_path)
    
    if column in df.columns:
        filtered = df[df[column] == value]
        write_excel(filtered, output)
        print(f"✓ 筛选结果已保存到: {output}")
    else:
        print(f"列 '{column}' 不存在")


def group_data(file_path: str, group_by: list, agg: dict, output: str):
    """分组统计"""
    df = read_excel(file_path)
    
    result = df.groupby(group_by).agg(agg)
    result.to_excel(output)
    print(f"✓ 分组统计已保存到: {output}")


def merge_files(file_paths: list, output: str, sheet: str = "Sheet1"):
    """合并多个 Excel 文件"""
    dfs = [read_excel(f) for f in file_paths]
    merged = pd.concat(dfs, ignore_index=True)
    write_excel(merged, output, sheet)
    print(f"✓ 已合并 {len(file_paths)} 个文件到: {output}")


def split_by_column(file_path: str, column: str, output_dir: str):
    """按列拆分 Excel"""
    df = read_excel(file_path)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    if column not in df.columns:
        print(f"列 '{column}' 不存在")
        return
    
    for value in df[column].unique():
        filtered = df[df[column] == value]
        safe_value = str(value).replace("/", "_").replace("\\", "_")
        output_file = output_path / f"{safe_value}.xlsx"
        write_excel(filtered, str(output_file))
    
    print(f"✓ 已按 '{column}' 拆分到: {output_dir}")


def create_template(output: str, template_type: str = "basic"):
    """创建模板"""
    wb = create_workbook()
    ws = wb.active
    
    if template_type == "basic":
        ws.title = "数据"
        headers = ["ID", "名称", "数量", "金额", "日期", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
    
    elif template_type == "report":
        ws.title = "报告"
        ws['A1'] = "报告标题"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A3'] = "日期："
        ws['B3'] = "2026-03-26"
        ws['A5'] = "摘要"
        ws['A6'] = "1. ..."
        ws['A7'] = "2. ..."
        ws['A9'] = "详细数据"
    
    elif template_type == "invoice":
        ws.title = "发票"
        ws['A1'] = "发票"
        ws['A1'].font = Font(bold=True, size=18)
        ws['A3'] = "发票编号："
        ws['C3'] = "日期："
        
        headers = ["项目", "描述", "数量", "单价", "金额"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        ws['D10'] = "合计："
        ws['E10'] = "=SUM(E6:E9)"
        ws['E10'].font = Font(bold=True)
    
    wb.save(output)
    print(f"✓ 模板已保存到: {output}")


def main():
    parser = argparse.ArgumentParser(description="Excel 表格处理工具")
    subparsers = parser.add_subparsers(dest="command", required=True)
    
    # 读取
    read_parser = subparsers.add_parser("read", help="读取 Excel")
    read_parser.add_argument("file", help="Excel 文件路径")
    read_parser.add_argument("-s", "--sheet", help="工作表名称")
    
    # 写入
    write_parser = subparsers.add_parser("write", help="写入 Excel")
    write_parser.add_argument("file", help="输出文件路径")
    write_parser.add_argument("-d", "--data", required=True, help="数据文件（CSV/JSON）")
    write_parser.add_argument("-s", "--sheet", default="Sheet1", help="工作表名称")
    
    # 分析
    analyze_parser = subparsers.add_parser("analyze", help="数据分析")
    analyze_parser.add_argument("file", help="Excel 文件路径")
    analyze_parser.add_argument("-s", "--sheet", help="工作表名称")
    
    # 筛选
    filter_parser = subparsers.add_parser("filter", help="筛选数据")
    filter_parser.add_argument("file", help="Excel 文件路径")
    filter_parser.add_argument("-c", "--column", required=True, help="筛选列")
    filter_parser.add_argument("-v", "--value", required=True, help="筛选值")
    filter_parser.add_argument("-o", "--output", required=True, help="输出文件")
    
    # 合并
    merge_parser = subparsers.add_parser("merge", help="合并文件")
    merge_parser.add_argument("files", nargs="+", help="Excel 文件列表")
    merge_parser.add_argument("-o", "--output", required=True, help="输出文件")
    
    # 拆分
    split_parser = subparsers.add_parser("split", help="拆分文件")
    split_parser.add_argument("file", help="Excel 文件路径")
    split_parser.add_argument("-c", "--column", required=True, help="拆分列")
    split_parser.add_argument("-o", "--output", required=True, help="输出目录")
    
    # 模板
    template_parser = subparsers.add_parser("template", help="创建模板")
    template_parser.add_argument("output", help="输出文件路径")
    template_parser.add_argument("-t", "--type", choices=["basic", "report", "invoice"],
                                  default="basic", help="模板类型")
    
    args = parser.parse_args()
    
    if args.command == "read":
        df = read_excel(args.file, args.sheet)
        print(df)
    
    elif args.command == "write":
        if args.data.endswith(".csv"):
            df = pd.read_csv(args.data)
        elif args.data.endswith(".json"):
            df = pd.read_json(args.data)
        else:
            df = pd.read_excel(args.data)
        write_excel(df, args.file, args.sheet)
    
    elif args.command == "analyze":
        analyze_data(args.file, args.sheet)
    
    elif args.command == "filter":
        filter_data(args.file, args.column, args.value, args.output)
    
    elif args.command == "merge":
        merge_files(args.files, args.output)
    
    elif args.command == "split":
        split_by_column(args.file, args.column, args.output)
    
    elif args.command == "template":
        create_template(args.output, args.type)


if __name__ == "__main__":
    main()